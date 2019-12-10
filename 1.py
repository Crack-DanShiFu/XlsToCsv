import hashlib
import os
import re

from openpyxl import Workbook
from openpyxl import load_workbook
from xml.dom.minidom import Document


def has_md5(s):
    m = hashlib.md5()  # 声明一个对象
    m.update(s.encode('UTF-8'))
    return m.hexdigest()


def load_data(path):
    wb = load_workbook(path)
    wb.guess_types = True  # 猜测格式类型
    ws = wb.active
    for i in ws['H'][1:]:
        pattern1 = re.compile(r'(\d+月\d+日)[,，]?(.*?)。', re.S)
        result = pattern1.findall(str(i.value))
        if len(result) == 0:
            pattern2 = re.compile(r'(.*?)。', re.S)
            flag = pattern2.findall(str(i.value))
            if len(flag) == 0:
                pattern3 = re.compile(r'(.*?)[，,]', re.S)
                flag = pattern3.findall(str(i.value))
            for res in flag:
                d = ws['E' + str(i.row)].value.split('-')
                result.append((str(int(d[1])) + '月' + str(int(d[2])) + '日', res))
        a = {
            'title': ws['B' + str(i.row)].value,
            'createDate': ws['C' + str(i.row)].value,
            'source': ws['D' + str(i.row)].value,
            'newsDate': ws['E' + str(i.row)].value,
            'author': ws['G' + str(i.row)].value,
            'eventDate': result[0][0],
            'event': result[0][1],
            'content': [snt.strip() for snt in ws['H' + str(i.row)].value.split('。')]
        }
        serial_num = has_md5(ws['B' + str(i.row)].value)
        write_xml(path, serial_num, a)


def write_xml(path, serial_num, source):
    doc = Document()  # 创建DOM文档对象
    DOCUMENT = doc.createElement('news')  # 创建根元素
    DOCUMENT.setAttribute('xmlns:xsi', "http://www.w3.org/2001/XMLSchema-instance")  # 设置命名空间
    DOCUMENT.setAttribute('xsi:noNamespaceSchemaLocation',
                          "file:///C:/Users/FXW/Desktop/201910_paper/newsSchema.xsd")  # 设置命名空间
    doc.appendChild(DOCUMENT)
    # 一级节点
    node1 = [
        doc.createElement('inf'),
        doc.createElement('evt'),
        doc.createElement('elements'),
        doc.createElement('standpoint'),
        doc.createElement('content')
    ]
    for n in node1:
        DOCUMENT.appendChild(n)
    #########
    # 二级节点
    ## inf
    node2_inf = [
        doc.createElement('title'),
        doc.createElement('createDate'),
        doc.createElement('source'),
        doc.createElement('newsDate'),
        doc.createElement('author'),
    ]
    ##evt
    node2_evt = [
        doc.createElement('eventDate'),
        doc.createElement('event'),
    ]
    ##elements
    node2_elements = [
        doc.createElement('gov'),
        doc.createElement('subject'),
        doc.createElement('media'),
        doc.createElement('field'),
    ]
    ##standpoint
    node2_standpoint = [
        doc.createElement('standwords'),
        doc.createElement('standscore'),
    ]
    #########写入个节点的参数

    node2_inf[0].appendChild(doc.createTextNode(source['title']))
    node2_inf[1].appendChild(doc.createTextNode(source['createDate']))
    node2_inf[2].appendChild(doc.createTextNode(source['source']))
    node2_inf[3].appendChild(doc.createTextNode(source['newsDate']))
    node2_inf[4].appendChild(doc.createTextNode(source['author']))
    node2_evt[0].appendChild(doc.createTextNode(source['eventDate']))
    node2_evt[1].appendChild(doc.createTextNode(source['event']))

    for s in source['content']:
        snt = doc.createElement('snt')
        snt.appendChild(doc.createTextNode(s))
        node1[-1].appendChild(snt)
    #########
    for n in node2_inf:
        node1[0].appendChild(n)
    for n in node2_evt:
        node1[1].appendChild(n)
    for n in node2_elements:
        node1[2].appendChild(n)
    for n in node2_standpoint:
        node1[3].appendChild(n)
    ###
    ########### 将DOM对象doc写入文件
    if not os.path.exists('result'):
        os.mkdir('result')
    save_path = 'result/' + path.split('/')[-1].split('.')[0]
    if not os.path.exists(save_path):
        os.mkdir(save_path)
    f = open(save_path + '/' + str(serial_num) + '.xml', 'w')
    doc.writexml(f, indent='\t', newl='\n', addindent='\t', encoding='utf-8')
    f.close()


if __name__ == '__main__':
    with open('1.txt', "r") as f:  # 设置文件对象
        path_list = f.readlines()
        for p in path_list:
            p = p.strip()
            load_data(p)
            print(p)
